import json
import docx
from openpyxl import load_workbook

# Replace 'your_file.xlsx' with the path to your Excel file.
workbook = load_workbook("lists.xlsx")
sheet = workbook.active

data = {}
first_data_point = ""
# You can also specify a specific sheet within the Excel workbook:
# excel_file = pd.read_excel('your_file.xlsx', sheet_name='Sheet1')

for column in sheet.iter_cols():
    # 'row' is a tuple representing one row of data.
    # You can access the columns by index, for example:
    i = 0
    for index, cell in enumerate(column):
        column_value = cell.value  # Access the cell value.
        if (i == 0):
            first_data_point = column_value
            i += 1
        else:
            if first_data_point in data:
                data[first_data_point].append(column_value)
            else:
                data[first_data_point] = [column_value]

print(data)
filename = "data.json"

with open(filename, "w") as json_file:
    json.dump(data, json_file, indent=4)
